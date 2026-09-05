from datetime import datetime, timedelta
import json
from openpyxl import load_workbook

source = '/Users/mac/Downloads/DATA H2H SISWA LPK YAMAGUCHI (1).xlsx'
wb = load_workbook(source, data_only=True, read_only=True)
rows = []
for year in (2024, 2025, 2026):
    ws = wb[str(year)]
    for values in ws.iter_rows(min_row=5, values_only=True):
        if not values[5] or not values[6]:
            continue
        def excel_date(value):
            if not value:
                return None
            if isinstance(value, datetime):
                return value.date().isoformat()
            if isinstance(value, (int, float)):
                return (datetime(1899, 12, 30) + timedelta(days=value)).date().isoformat()
            return None
        rows.append({
            'nis': str(values[5]).strip(), 'full_name': str(values[6]).strip(),
            'school_name': str(values[7]).strip() if values[7] else None,
            'enrollment_date': excel_date(values[8]), 'graduation_date': excel_date(values[9]),
            'departure_date': excel_date(values[10]), 'job_sector': values[11],
            'placement': values[12], 'notes': values[13],
            'status': 'keluar' if str(values[13]).strip().upper() == 'KELUAR' else ('pending' if str(values[13]).strip().upper() == 'PENDING' else 'aktif'),
        })
with open('storage/app/h2h_students.json', 'w') as f:
    json.dump(rows, f, ensure_ascii=False)
print(len(rows))
